#!/usr/bin/env python3
"""
RESEARCH QUEUE — scans hull workbooks for [TBC] / [MISSING] / blank critical
cells and produces a prioritised research list with ready-made search queries.
Usage:  python research_queue.py <workbook1.xlsx> [workbook2.xlsx ...]

Per CLAUDE.md rules: unknowns are NEVER filled by guesswork. This tool
collects every unknown into Research_Queue.xlsx so each can be resolved by
web search (cite source) or supplier RFI — and flagged if unresolvable.

Output columns:
  Workbook | Sheet | Row | Item | Equipment | Problem | Suggested Search Query | Status
"""
import re, sys
import openpyxl
from openpyxl.styles import Font

FLAG = re.compile(r"\[(TBC|MISSING|CA-\d+)\]", re.I)

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def main(paths):
    ob = openpyxl.Workbook(); out = ob.active; out.title = "RESEARCH QUEUE"
    out.append(["RESEARCH QUEUE — unresolved items needing web search / supplier RFI"])
    out.append(["Rule: resolve with cited source, or mark UNRESOLVABLE and escalate. Never guess."])
    out.append([])
    out.append(["Workbook","Sheet","Row","Item","Equipment","Problem","Suggested Search Query","Status"])
    for c in out[4]: c.font = Font(bold=True)
    n = 0
    for path in paths:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            out.append([path, "", "", "", "", f"[ERROR] cannot open: {e}", "", "OPEN"]); n += 1; continue
        short = path.split("/")[-1].split("\\")[-1]
        for ws in wb.worksheets:
            for ri, r in enumerate(ws.iter_rows(values_only=True), 1):
                a = str(r[0]) if r[0] is not None else ""
                isdata = bool(re.match(r"^\d+\.\d+", a))
                rowtxt = " | ".join(str(c) for c in r if c is not None)
                eq = str(r[1])[:45] if len(r) > 1 and r[1] else ""
                mk = str(r[2])[:45] if len(r) > 2 and r[2] else ""
                problems = []
                for m in FLAG.finditer(rowtxt):
                    problems.append(f"[{m.group(1).upper()}] tag present")
                if isdata and "LOAD LIST" in ws.title.upper() and len(r) >= 16:
                    qty, kw, I = num(r[3]), num(r[7]), num(r[8])
                    if qty and qty > 0:
                        if kw is None: problems.append("Power kW blank")
                        if kw and kw > 0 and I is None: problems.append("Current A blank")
                        if r[15] is None or str(r[15]).strip() == "": problems.append("Remarks blank (rule: cite source)")
                if not problems: continue
                q = f"{mk or eq} marine datasheet power rating specifications" if (mk or eq) else ""
                out.append([short, ws.title, ri, a, eq or mk, "; ".join(sorted(set(problems))), q, "OPEN"])
                n += 1
        wb.close()
    out.append([]); out.append([f"TOTAL OPEN ITEMS: {n}"])
    for col, w in zip("ABCDEFGH", (40, 22, 6, 8, 45, 55, 55, 10)):
        out.column_dimensions[col].width = w
    ob.save("Research_Queue.xlsx")
    print(f"✓ Research_Queue.xlsx: {n} open items from {len(paths)} workbook(s)")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    main(sys.argv[1:])
