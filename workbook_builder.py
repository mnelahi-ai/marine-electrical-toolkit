#!/usr/bin/env python3
"""
WORKBOOK BUILDER — creates a Load List + Alarm List workbook skeleton
following the H674 template rules (D1/D2/D3 of the audit prompt).
Usage:  python workbook_builder.py hull_config.json [output.xlsx]

hull_config.json example:
{
  "hull": "H699", "rev": "Rev 1", "date": "10/06/2026", "prepared_by": "Mashuk",
  "contractor": "Z-POWER AUTOMATION PTE LTD | STRATEGIC MARINE (S) PTE LTD",
  "vessel_type": "42.0 M Fast Crew Boat", "classification": "LR 100A1 SSC WORKBOAT MONO G4",
  "voltage_system": "415/240 VAC 3PH 4W 50Hz + 24 VDC",
  "main_engine": "...", "gearbox": "...", "bow_thruster": "...", "generators": "...",
  "voltage_separator": "~"
}
All unspecified fields are written as [TBC] — NEVER guessed.
Current formula locked: 3PH I=kW*1000/(1.732*V*0.85); 1PH PF 0.85; DC I=kW*1000/V.
"""
import json, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

SECTIONS = [
    "1.0  POWER GENERATION & DISTRIBUTION", "2.0  BATTERIES & CHARGERS",
    "3.0  PROPULSION, GEARBOX, BOW THRUSTER & STEERING", "4.0  DECK & MACHINERY",
    "5.0  FIRE-FIGHTING & SAFETY", "6.0  PUMPING SYSTEMS",
    "7.0  HVAC & ACCOMMODATION", "8.0  LIGHTING, GPO & GENERAL",
    "9.0  NAVIGATION & COMMUNICATION (GMDSS A2)", "10.0  MONITORING & CONTROL (AMS)",
    "11.0  ENTERTAINMENT & CREW WELFARE",
]
ALARM_GROUPS = [
    "1  FIRE / GAS / SAFETY", "2  BILGE HIGH LEVEL (P1 — every WT space, LR SSC Pt.5 Ch.2 §2.3)",
    "3  WT / WEATHERTIGHT DOORS & ESCAPE HATCHES", "4  GENERATOR / MSB / EARTH FAULTS",
    "5  BATTERIES & CHARGERS", "6  VENT FANS / PUMPS / MACHINERY STATUS",
    "7  TANK LEVEL GAUGING", "8  NAV LIGHTS / COMMS / MISC", "9  ADDED ALARMS",
]
BOLD = Font(bold=True); BANNER = PatternFill("solid", fgColor="1F4E79")
BANNER_F = Font(bold=True, color="FFFFFF")

def g(cfg, key): return cfg.get(key) or "[TBC]"

def build(cfg_path, out=None):
    cfg = json.load(open(cfg_path))
    hull = g(cfg, "hull"); rev = cfg.get("rev", "Rev 1"); date = g(cfg, "date")
    out = out or f"{hull}_Load List_Alarm_Point_List_{rev.replace(' ','')}.xlsx"
    wb = openpyxl.Workbook()

    # ── Sheet 1 COVER (D1) ──
    cv = wb.active; cv.title = "COVER"
    rows = [
        (g(cfg,"hull_tagline") if cfg.get("hull_tagline") else hull, rev),
        (g(cfg,"contractor"), "STRATEGIC MARINE (S) PTE LTD"),
        ("TECHNICAL SPECIFICATION", f"{hull} ELECTRICAL LOAD LIST + AMS I/O LIST"),
        ("Project Number", g(cfg,"project_number")),
        ("Drawing Ref (Primary)", g(cfg,"drawing_primary")),
        ("Drawing Ref (Lighting/GPO)", g(cfg,"drawing_lighting")),
        ("Drawing Ref (HVAC)", g(cfg,"drawing_hvac")),
        ("Vessel Type", g(cfg,"vessel_type")),
        ("Hull Numbers", g(cfg,"hull_numbers") if cfg.get("hull_numbers") else hull),
        ("Classification", g(cfg,"classification")),
        ("Voltage System", g(cfg,"voltage_system")),
        ("Main Engine", g(cfg,"main_engine")),
        ("Gearbox", g(cfg,"gearbox")),
        ("Bow Thruster", g(cfg,"bow_thruster")),
        ("Generators", g(cfg,"generators")),
        ("Shore Supply", g(cfg,"shore_supply")),
        ("SC Level", g(cfg,"sc_level")),
        ("Cable Standard", g(cfg,"cable_standard")),
        ("Report Rev", f"{rev} — Initial issue for {hull}"),
        ("Date", date),
        ("Prepared by", g(cfg,"prepared_by")),
        ("Approved by", "—"),
    ]
    for a, b in rows: cv.append([a, b])
    cv.append([])
    cv.append(["CHANGE LOG — Change No. | Date | Section | Description | Status"])
    cv["A24"].font = BOLD
    for r in range(25, 35): cv.append([])
    cv.append([]); cv.append(["WORKBOOK CONTENTS"])
    cv.append(["Sheet 1", "COVER"]); cv.append(["Sheet 2", f"{hull} LOAD LIST"])
    cv.append(["Sheet 3", "ALARM LIST"])
    cv.column_dimensions["A"].width = 30; cv.column_dimensions["B"].width = 70
    for r in (1, 2, 3): cv.cell(r, 1).font = BOLD

    # ── Sheet 2 LOAD LIST (D2) ──
    ll = wb.create_sheet(f"{hull} LOAD LIST")
    ll.append([f"{hull} ELECTRICAL LOAD LIST — {rev}"])
    ll.append([f"Refs: {g(cfg,'drawing_primary')}"])
    ll.append(["Current formula (PF=0.85): 3PH I=kW×1000/(√3×V×0.85) | 1PH I=kW×1000/(V×0.85) | DC I=kW×1000/V. "
               "Worked: 3PH 15kW 415V=24.56A | 1PH 0.75kW 240V=3.68A | DC 0.26kW 24V=10.83A"])
    ll.append(["Item\nNo.","Equipment / System Name","Make / Model / Specification","QTY","VOLTAGE\n(V)","PHASE",
               "FREQ\n(Hz)","POWER\n(kW)","CURRENT\n(A)","TOTAL\nLOAD (kW)","Electrical SLD & Termination",
               "Cable Schedule","Supplier / Provider","Date Recd","Wt (kg)","Remarks / Audit Tag"])
    for c in ll[4]: c.font = BOLD
    ll["A1"].font = Font(bold=True, size=13)
    r = 5
    for s in SECTIONS:
        ll.cell(r, 1, f"▌ {s}")
        for col in range(1, 17):
            ll.cell(r, col).fill = BANNER
        ll.cell(r, 1).font = BANNER_F
        r += 1
        n = s.split()[0]
        ll.cell(r, 1, f"{n.replace('.0','')}" + ".1"); ll.cell(r, 2, "[TBC] — populate from source documents")
        ll.cell(r, 16, "[TBC] — every row must cite its source doc"); r += 1
        ll.cell(r, 1, f"{s} TOTAL"); ll.cell(r, 1).font = BOLD; ll.cell(r, 10, 0); r += 1
    ll.cell(r, 1, "GRAND TOTAL (SUM OF SECTION SUBTOTALS)"); ll.cell(r, 1).font = BOLD
    widths = (8,36,38,6,11,8,8,9,10,12,26,22,20,12,9,44)
    for col, w in zip("ABCDEFGHIJKLMNOP", widths): ll.column_dimensions[col].width = w

    # ── Sheet 3 ALARM LIST (D3) ──
    al = wb.create_sheet("ALARM LIST")
    al.append([f"{hull} — ALARM MONITORING SYSTEM (AMS) I/O GAP ANALYSIS"])
    std = cfg.get("alarm_standards", "LR SSC 2024 | IEC 62933")
    al.append([f"AMS Panel: {g(cfg,'ams_panel')} | I/O doc: {g(cfg,'ams_io_doc')} | Standards: {std}"])
    al.append(["Legend: 🟢 PRESENT | 🔴 MISSING | 🟡 REVIEW | P1 Safety / P2 Class / P3 Good practice"])
    al.append([]); al.append([])
    al.append(["No","Description","Condition","LR/IEC Req?","Rule ref","Priority"])
    for c in al[6]: c.font = BOLD
    al["A1"].font = Font(bold=True, size=13)
    r = 7
    for gname in ALARM_GROUPS:
        al.cell(r, 1, f"▌ GROUP {gname}")
        for col in range(1, 7): al.cell(r, col).fill = BANNER
        al.cell(r, 1).font = BANNER_F; r += 1
        al.cell(r, 2, "[TBC] — populate from AMS I/O list / class rules"); r += 1
    for col, w in zip("ABCDEF", (7, 52, 26, 12, 30, 9)): al.column_dimensions[col].width = w

    wb.save(out)
    print(f"✓ {out} created — all unknown fields tagged [TBC]; populate from sources, never guess.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    build(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
