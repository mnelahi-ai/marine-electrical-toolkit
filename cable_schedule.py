#!/usr/bin/env python3
"""
CABLE SCHEDULE GENERATOR — reads a hull Load List workbook and produces a
Cable Schedule with suggested conductor sizes.
Usage:  python cable_schedule.py <LoadList.xlsx> [output.xlsx]

For every load-list data row with QTY>0 and kW>0:
  - takes Current (A) from the workbook (or recomputes at PF 0.85 if blank)
  - suggests minimum conductor size from a marine cable current-rating table
    (IEC 60092-352 style, XLPE 90°C, bunched, 45°C ambient — INDICATIVE)
  - cable LENGTH is written as [TBC] — volt-drop check must be completed
    once route lengths are measured. NO LENGTH IS EVER ASSUMED.

Output columns:
  Cable No | From | To (Equipment) | V | Phase | I (A) | Cores | Size mm² |
  Type | Length m | VD check | Remarks

VERIFY all sizes against the project cable type approvals (Cable folder)
and class rules before issue. This tool proposes; the engineer disposes.
"""
import math, re, sys
import openpyxl
from openpyxl.styles import Font

# Indicative continuous ratings (A), XLPE 90°C marine cable, 45°C ambient.
# (size mm², 3-4 core rating). VERIFY against manufacturer/class tables.
RATING = [(1.5,14),(2.5,20),(4,27),(6,34),(10,48),(16,64),(25,85),(35,105),
          (50,130),(70,165),(95,200),(120,230),(150,265),(185,300),(240,355),(300,405)]

def size_for(I):
    for s, amp in RATING:
        if amp >= I: return s
    return "PARALLEL RUNS [TBC]"

def num(x):
    try: return float(x)
    except (TypeError, ValueError): return None

def main(src, out=None):
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames if "LOAD LIST" in s.upper()), wb.sheetnames[0])
    ws = wb[sheet]
    hull = sheet.replace(" LOAD LIST", "")
    out = out or f"{hull}_Cable_Schedule_Rev0.xlsx"
    ob = openpyxl.Workbook(); os_ = ob.active; os_.title = f"{hull} CABLE SCHEDULE"
    os_.append([f"{hull} CABLE SCHEDULE — Rev 0 (generated from {src.split('/')[-1]})"])
    os_.append(["INDICATIVE ONLY — sizes from generic IEC 60092-352-style table. "
                "Verify vs project cable type approvals + class rules. Lengths [TBC] — measure routes."])
    os_.append([])
    os_.append(["Cable No","From","To (Equipment)","V","Phase","I (A)","Cores","Min Size mm²",
                "Type","Length m","VD Check","Remarks"])
    for c in os_[4]: c.font = Font(bold=True)
    n = 0; sec = "0"
    for r in ws.iter_rows(values_only=True):
        a = str(r[0]) if r[0] is not None else ""
        if a.startswith("▌"):
            m = re.search(r"(\d+)\.0", a); sec = m.group(1) if m else sec
            continue
        if not re.match(r"^\d+\.\d+", a): continue
        eq, qty, V, ph, kw, I = r[1], num(r[3]), str(r[4] or ""), str(r[5] or ""), num(r[7]), num(r[8])
        fr = r[10] or "[TBC]"
        if not qty or not kw or kw <= 0: continue
        mv = re.findall(r"\d+", V); vn = float(mv[-1]) if mv else None
        note = ""
        if I is None and vn:
            I = kw*1000/(math.sqrt(3)*vn*0.85) if "3" in ph else (kw*1000/vn if "DC" in ph.upper() else kw*1000/(vn*0.85))
            note = "I recomputed (PF 0.85); "
        if I is None:
            os_.append([f"{hull}-C{sec}.{n+1:02d}", fr, eq, V, ph, "[MISSING]", "", "", "", "[TBC]", "[TBC]",
                        "[MISSING] no current & no voltage — resolve source data first"])
            n += 1; continue
        cores = "2C" if "DC" in ph.upper() or ("1" in ph and "3" not in ph) else "4C(3C+E)"
        os_.append([f"{hull}-C{sec}.{n+1:02d}", fr, eq, V, ph, round(I, 1), cores, size_for(I),
                    "XLPE/SHF1 marine [verify approval]", "[TBC]", "[TBC] after route length",
                    note + f"src row {a}; size INDICATIVE"])
        n += 1
    os_.append([]); os_.append([f"TOTAL CABLES: {n}"])
    for col, w in zip("ABCDEFGHIJKL", (13,26,38,11,7,9,9,13,30,9,18,46)):
        os_.column_dimensions[col].width = w
    ob.save(out); wb.close()
    print(f"✓ {out}: {n} cables generated from {sheet}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
